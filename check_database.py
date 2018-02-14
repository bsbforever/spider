#coding=utf8


from ping3 import ping, verbose_ping
import socket
from openpyxl import Workbook
from openpyxl import load_workbook


def ifping(ipaddress):
    result=ping( ipaddress)
    return result

def  checkport(ipaddress,port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex((ipaddress, port))
    return result




if __name__=="__main__":
    #ipaddresses=['10.65.206.23','10.65.1.168','10.65.202.84']
    dbs=['1521:oracle','1522:oracle','1523:oracle','1525:oracle','1527:oracle','1529:oracle','1433:sqlserver','3306:mysql']
    networks=['10.65.9','10.65.12','10.65.31','10.65.34','10.65.203','10.65.180','10.65.255','10.65.3','10.65.4','10.65.5','10.65.6','10.65.8','10.65.10'
        , '10.65.27','10.65.29','10.65.35','10.65.36','10.65.45','10.65.46','10.65.47','10.65.48','10.65.49','10.65.184','192.168.40',
              '10.65.14','10.65.15','10.65.16','10.65.24','10.65.182','192.168.50','10.65.2','10.65.22','10.65.25','10.65.26','10.65.181'
        , '10.65.11','10.65.41','10.65.42','10.65.43','10.65.44','10.65.184','10.65.13','10.65.17','10.65.21','10.65.23','10.65.201','10.65.202','10.65.38'
           , '10.60.14','130.147.192']
    path=r'X:\Users\42274\Desktop\excel\dbscan.xlsx'
    wb = load_workbook(path)

    # 初始化第一个sheet
    #work_sheet = wb.active
    #work_sheet.title = "range names"
    for network in networks:
        ## 新建sheet用于存放各个网段并初始化
        if  network in wb.sheetnames:
            work_sheet = wb.get_sheet_by_name(network)
        else:
            wb.create_sheet(title=network)
            work_sheet = wb.get_sheet_by_name(network)
            work_sheet.append(['IP地址', '数据库'])
            wb.save(path)

        for i in range(1,255):
            ipaddress=network+'.'+str(i)
            print ('now checking '+ipaddress)
            result = ifping(ipaddress)
            if result is  None:
                #print(ipaddress + ' is  not reachable')
                work_sheet.append([ipaddress, 'not reachable'])
                wb.save(path)

            else:
                #print(ipaddress + ' is  reachable')
                for db in dbs:
                    opened=0                        #用于判断是否无DB端口开放
                    port=int(db.split(':')[0])
                    name=db.split(':')[1]
                    ifopen=checkport(ipaddress, port)
                    #print (result)
                    if ifopen==0:
                        #print (ipaddress +' have '+name +' installed')
                        opened=opened+1
                        work_sheet.append([ipaddress,name])
                        wb.save(path)
                if opened==0:
                    work_sheet.append([ipaddress, 'No DBs Found'])
                    wb.save(path)
    #wb.save(path)
    wb.close()


